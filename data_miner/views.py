from django.shortcuts import render
from django.views.generic import TemplateView
from django.http import JsonResponse, HttpResponse, FileResponse
from .scrapper import EnhancedContactScraper
from .models import MiningHistory
import pandas as pd
import logging
import os
from django.conf import settings
import uuid
from datetime import datetime
import json
from openpyxl import Workbook
import asyncio
import time

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Create your views here.
class DataMinerView(TemplateView):
    template_name = 'data miner/miner.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Data Mining Tool'
        context['app_name'] = 'Data Mining Tool'
        
        # Get recent mining history for the current user
        if self.request.user.is_authenticated:
            recent_minings = MiningHistory.objects.filter(user=self.request.user).order_by('-created_at')[:5]
            context['recent_minings'] = recent_minings

        # If this is a GET request with search parameters, perform the search
        if self.request.method == 'GET' and 'keyword' in self.request.GET:
            try:
                keyword = self.request.GET.get('keyword', '').strip()
                data_type = self.request.GET.get('data_type', 'phone')
                country = self.request.GET.get('country', 'IN')
                
                if keyword and len(keyword) >= 3:
                    # Initialize scraper
                    scraper = EnhancedContactScraper()
                    results = scraper.search_and_extract(
                        target=keyword,
                        country=country,
                        max_results=30,
                        exact_count=False,
                        max_pages=3
                    )
                    
                    if results and not isinstance(results, dict) or ('error' not in results):
                        context['initial_results'] = {
                            'success': True,
                            'data': results,
                            'message': f"Found {len(results.get('phones', [])) if data_type == 'phone' else len(results.get('emails', []))} results"
                        }
                    else:
                        context['initial_results'] = {
                            'error': results.get('error', 'No results found')
                        }
            except Exception as e:
                logger.error(f"Error in GET request search: {e}")
                context['initial_results'] = {
                    'error': 'An error occurred during the search'
                }
            finally:
                if 'scraper' in locals():
                    try:
                        scraper.close_browser()
                    except:
                        pass
        
        return context

    def generate_excel(self, results, keyword, data_type):
        # Create a DataFrame
        df = pd.DataFrame(results, columns=[data_type.capitalize()])
        
        # Generate unique filename
        filename = f"mining_results_{keyword.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(settings.MEDIA_ROOT, 'mining_results', filename)
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Save to Excel
        df.to_excel(filepath, index=False)
        
        # Return relative path from MEDIA_ROOT
        return os.path.join('mining_results', filename)

    def post(self, request, *args, **kwargs):
        # Check if it's a download request
        if request.headers.get('Content-Type') == 'application/json':
            try:
                data = json.loads(request.body)
                results = data.get('results', [])
                keyword = data.get('keyword', '')
                data_type = data.get('data_type', '')
                
                if not results:
                    return JsonResponse({'error': 'No results to download'}, status=400)
                
                # Create Excel workbook
                wb = Workbook()
                ws = wb.active
                ws.title = 'Mining Results'
                
                # Add headers
                ws['A1'] = data_type.capitalize()
                
                # Add data
                for i, result in enumerate(results, start=2):
                    ws[f'A{i}'] = result
                
                # Save to a temporary file
                temp_filename = f"temp_mining_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                temp_filepath = os.path.join(settings.MEDIA_ROOT, 'temp', temp_filename)
                os.makedirs(os.path.dirname(temp_filepath), exist_ok=True)
                wb.save(temp_filepath)
                
                # Return file as attachment
                with open(temp_filepath, 'rb') as excel_file:
                    response = HttpResponse(
                        excel_file.read(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = f'attachment; filename=mining_results_{keyword.replace(" ", "_")}.xlsx'
                
                # Clean up temporary file
                try:
                    os.remove(temp_filepath)
                except:
                    pass
                    
                return response
                
            except Exception as e:
                logger.error(f"Error generating Excel file: {e}")
                return JsonResponse({'error': 'Failed to generate Excel file'}, status=500)

        # If not a download request, proceed with the original post method
        try:
            # Get and validate input parameters
            keyword = request.POST.get('keyword', '').strip()
            data_type = request.POST.get('data_type', 'phone')  # 'phone' or 'email'
            country = request.POST.get('country', 'IN')  # Default to India
            max_results = 30  # Fixed to 30 results

            # Validate keyword
            if not keyword:
                return JsonResponse({'error': 'Please enter a search keyword'}, status=400)
            
            if len(keyword) < 3:
                return JsonResponse({'error': 'Search keyword must be at least 3 characters long'}, status=400)

            # Validate data type
            if data_type not in ['phone', 'email']:
                return JsonResponse({'error': 'Invalid data type specified'}, status=400)

            # Check if we already have recent results for this search
            if request.user.is_authenticated:
                recent_mining = MiningHistory.objects.filter(
                    user=request.user,
                    keyword=keyword,
                    country=country,
                    data_type=data_type,
                    created_at__gte=datetime.now().replace(hour=0, minute=0, second=0)  # From today
                ).first()
                
                if recent_mining:
                    # Return cached results
                    try:
                        with open(recent_mining.excel_file.path, 'r') as f:
                            df = pd.read_excel(recent_mining.excel_file.path)
                            results = df[data_type.capitalize()].tolist()
                            return JsonResponse({
                                'success': True,
                                'data': {data_type + 's': results},
                                'message': f"Found {len(results)} results (cached)",
                                'elapsed_time': 0
                            })
                    except:
                        pass

            # Initialize scraper with proper error handling
            try:
                scraper = EnhancedContactScraper()
                logger.info(f"Initialized scraper for keyword: {keyword}, data type: {data_type}")
            except Exception as e:
                logger.error(f"Failed to initialize scraper: {e}")
                return JsonResponse({'error': 'Failed to initialize search. Please try again.'}, status=500)
            
            try:
                # Start time for scraping
                start_time = time.time()

                # First attempt with fewer pages for quick results
                logger.info(f"Starting initial search for {keyword} with data type {data_type}")
                results = scraper.search_and_extract(
                    target=keyword,
                    country=country,
                    max_results=max_results,
                    exact_count=False,
                    max_pages=3
                )

                if not results:
                    return JsonResponse({
                        'error': 'No results found. Please try a different search term.',
                        'elapsed_time': time.time() - start_time
                    }, status=404)

                if isinstance(results, dict) and 'error' in results:
                    logger.error(f"Search error: {results['error']}")
                    return JsonResponse({
                        'error': results['error'],
                        'elapsed_time': time.time() - start_time
                    }, status=500)

                # Filter and prepare results based on selected data type
                filtered_results = {}
                final_results = []
                
                if data_type == 'phone':
                    phones = results.get('phones', [])
                    filtered_phones = [p for p in phones if p and len(p.strip()) >= 10][:max_results]
                    filtered_results['phones'] = filtered_phones
                    final_results = filtered_phones
                    
                if data_type == 'email':
                    emails = results.get('emails', [])
                    filtered_emails = [e for e in emails if e and '@' in e][:max_results]
                    filtered_results['emails'] = filtered_emails
                    final_results = filtered_emails

                # If we have enough results, don't do additional search
                if len(final_results) >= max_results:
                    logger.info(f"Found enough results ({len(final_results)}), skipping additional search")
                else:
                    # If we need more results and initial search was successful, try additional search
                    logger.info(f"Attempting additional search for more results")
                    additional_results = scraper.search_and_extract(
                        target=keyword,
                        country=country,
                        max_results=max_results,
                        exact_count=True,
                        max_pages=10
                    )
                    
                    if additional_results and isinstance(additional_results, dict):
                        if data_type == 'phone' and 'phones' in additional_results:
                            additional_phones = additional_results.get('phones', [])
                            filtered_results['phones'] = list(set(filtered_results.get('phones', []) + additional_phones))[:max_results]
                            final_results = filtered_results['phones']
                        if data_type == 'email' and 'emails' in additional_results:
                            additional_emails = additional_results.get('emails', [])
                            filtered_results['emails'] = list(set(filtered_results.get('emails', []) + additional_emails))[:max_results]
                            final_results = filtered_results['emails']

                # Calculate elapsed time
                elapsed_time = time.time() - start_time

                # Ensure we have at least some results
                if not final_results:
                    return JsonResponse({
                        'warning': 'No valid results found. Try broadening your search.',
                        'data': filtered_results,
                        'message': "No results found",
                        'elapsed_time': elapsed_time
                    }, status=200)

                # Generate Excel file and save to database if user is authenticated
                if request.user.is_authenticated and final_results:
                    try:
                        excel_path = self.generate_excel(final_results, keyword, data_type)
                        
                        # Save to database
                        MiningHistory.objects.create(
                            user=request.user,
                            keyword=keyword,
                            country=country,
                            data_type=data_type,
                            results_count=len(final_results),
                            excel_file=excel_path
                        )
                    except Exception as e:
                        logger.error(f"Error saving mining history: {str(e)}")
                        # Continue with the response even if saving fails
                
                return JsonResponse({
                    'success': True,
                    'data': filtered_results,
                    'message': f"Found {len(final_results)} results",
                    'elapsed_time': elapsed_time
                })

            except Exception as e:
                logger.error(f"Error during search and extract: {str(e)}")
                return JsonResponse({
                    'error': 'An error occurred during the search. Please try again.',
                    'elapsed_time': time.time() - start_time
                }, status=500)

        except Exception as e:
            logger.error(f"Unexpected error in post method: {str(e)}")
            return JsonResponse({
                'error': 'An unexpected error occurred. Please try again.',
                'elapsed_time': 0
            }, status=500)
        finally:
            # Ensure scraper resources are cleaned up
            if 'scraper' in locals():
                try:
                    scraper.close_browser()
                except:
                    pass

    def get_excel(self, request, history_id):
        try:
            history = MiningHistory.objects.get(id=history_id, user=request.user)
            if history.excel_file:
                return FileResponse(
                    open(history.excel_file.path, 'rb'),
                    as_attachment=True,
                    filename=f'mining_results_{history.keyword}.xlsx'
                )
        except MiningHistory.DoesNotExist:
            return JsonResponse({'error': 'File not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
